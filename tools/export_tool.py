# tools/export_tool.py

"""
ExportTool
Generates Excel files per expense category.
One file per category — new invoices are appended.
"""

import logging
import shutil
from collections import Counter, defaultdict
from pathlib import Path
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from config.settings import OUTPUTS_DIR, NEED_REVIEW_DIR
from models.invoice import Invoice

logger = logging.getLogger(__name__)

HEADER_COLOR = "1F3864"

CATEGORY_FILENAMES = {
    "SaaS":                      "SaaS",
    "Travel":                    "Travel",
    "Office":                    "Office",
    "Utilities":                 "Utilities",
    "IT Hardware":               "IT_Hardware",
    "Consulting & Services":     "Consulting_Services",
    "Marketing & Communication": "Marketing_Communication",
    "Legal & Compliance":        "Legal_Compliance",
    "Maintenance & Repair":      "Maintenance_Repair",
    "Other":                     "Other",
}
class ExportTool:

    def export(self, invoices: list) -> list:
        """Export validated invoices to Excel files per category."""
        if not invoices:
            logger.info("No invoices to export.")
            return []

        valides   = [inv for inv in invoices if inv.is_valid]
        invalides = len(invoices) - len(valides)
        if invalides > 0:
            logger.warning(
                f"{invalides} invalid invoice(s) excluded from export"
            )

        if not valides:
            logger.info("No valid invoices to export.")
            return []

        logger.info(f"Exporting {len(valides)} invoice(s)...")

        grouped = defaultdict(list)
        for invoice in valides:
            category = invoice.expense_category or "Other"
            grouped[category].append(invoice)

        generated_files = []
        for category, cat_invoices in grouped.items():
            path = self._export_category(category, cat_invoices)
            if path:
                generated_files.append(path)

        logger.info(
            f"Export complete: {len(generated_files)} file(s) generated."
        )
        return generated_files
    def _deduplicate_rows(self, existing_rows: list, new_rows: list) -> list:
        """
        Remove duplicate invoices based on Invoice ID + Vendor.
        Keeps the first occurrence.
        """
        seen = set()
        unique_rows = []
    
        for row in existing_rows + new_rows:
            key = (
                str(row.get("Invoice ID", "")).strip().lower(),
                str(row.get("Vendor", "")).strip().lower(),
        )
            if key not in seen:
                seen.add(key)
                unique_rows.append(row)
    
        return unique_rows
    def _export_category(self, category: str, invoices: list) -> Path | None:
        """Generate or update an Excel file for a category."""
        safe_name   = CATEGORY_FILENAMES.get(category, "Other")
        filename    = f"{safe_name}.xlsx"
        output_path = OUTPUTS_DIR / filename

        try:
            existing_rows = []
            if output_path.exists():
                try:
                    existing_df = pd.read_excel(output_path, sheet_name="Invoices")
                    existing_rows = existing_df.to_dict("records")
                    logger.info(
                        f"  [{category}] Appending {len(invoices)} "
                        f"to {len(existing_rows)} existing"
                    )
                except Exception as e:
                    logger.warning(
                        f"Cannot read existing file: {e}. Creating new file."
                    )
                    existing_rows = []

            new_rows = [inv.to_dict() for inv in invoices]
            all_rows = self._deduplicate_rows(existing_rows,new_rows)
            df = pd.DataFrame(all_rows)

            # Format numeric columns
            for col in ["Subtotal", "Tax", "Total", "Discount",
                        "Amount Due", "Previous Unpaid Balance", "DiT Score"]:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").round(3)

            all_invoices = self._rows_to_invoices(existing_rows) + invoices
            summary_df = self._build_summary(all_invoices)

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Invoices")
                self._format_sheet(writer.sheets["Invoices"], df)

                summary_df.to_excel(writer, index=False, sheet_name="Summary")
                self._format_sheet(writer.sheets["Summary"], summary_df)

            logger.info(
                f"  [{category}] → {filename} "
                f"({len(all_rows)} invoice(s) total)"
            )
            return output_path

        except Exception as e:
            logger.error(f"Export failed for '{category}': {e}")
            return None

    def _rows_to_invoices(self, rows: list) -> list:
        """Reconvert Excel rows to Invoice objects (pour le Summary)."""
        invoices = []
        for row in rows:
            inv = Invoice()
            inv.vendor_name               = row.get("Vendor")
            inv.vendor_address            = row.get("Vendor Address")
            inv.vendor_tax_id             = row.get("Vendor Tax ID")
            inv.customer_name             = row.get("Customer")
            inv.customer_id               = row.get("Customer ID")
            inv.customer_tax_id           = row.get("Customer Tax ID")
            inv.invoice_id                = row.get("Invoice ID")
            inv.purchase_order            = row.get("Purchase Order")
            inv.kvk_number                = row.get("KVK Number")
            inv.payment_term              = row.get("Payment Term")
            inv.invoice_date              = row.get("Invoice Date")
            inv.due_date                  = row.get("Due Date")
            inv.service_start_date        = row.get("Service Start Date")
            inv.service_end_date          = row.get("Service End Date")
            inv.total_amount              = row.get("Total")
            inv.subtotal                  = row.get("Subtotal")
            inv.tax_amount                = row.get("Tax")
            inv.total_discount            = row.get("Discount")
            inv.amount_due                = row.get("Amount Due")
            inv.previous_unpaid_balance   = row.get("Previous Unpaid Balance")
            inv.currency                  = row.get("Currency", "MAD")
            inv.billing_address           = row.get("Billing Address")
            inv.shipping_address          = row.get("Shipping Address")
            inv.expense_category          = row.get("Category")
            inv.llm_confidence            = row.get("LLM Confidence")
            dit = row.get("DiT Score")
            inv.dit_confidence            = float(dit) if dit else None
            invoices.append(inv)
        return invoices

    def _build_summary(self, invoices: list) -> pd.DataFrame:
        """Build Summary sheet with per-vendor statistics."""
        vendors = defaultdict(list)
        for inv in invoices:
            vendor = inv.vendor_name or "Unknown"
            vendors[vendor].append(inv)

        rows = []
        for vendor, vendor_invoices in vendors.items():
            amounts = [
                float(inv.total_amount)
                for inv in vendor_invoices
                if inv.total_amount is not None
            ]

            currencies = [
                inv.currency
                for inv in vendor_invoices
                if inv.currency
            ]
            currency = (
                Counter(currencies).most_common(1)[0][0]
                if currencies else "N/A"
            )

            rows.append({
                "Vendor":          vendor,
                "Invoice Count":   len(vendor_invoices),
                "Total Amount":    round(sum(amounts), 2),
                "Average Amount":  round(sum(amounts) / len(amounts), 2) if amounts else 0,
                "Min Amount":      round(min(amounts), 2) if amounts else 0,
                "Max Amount":      round(max(amounts), 2) if amounts else 0,
                "Currency":        currency,
            })

        rows.sort(key=lambda x: x["Total Amount"], reverse=True)
        return pd.DataFrame(rows)

    def _format_sheet(self, ws, df: pd.DataFrame) -> None:
        """Apply professional formatting to Excel sheet."""
        header_font  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        header_fill  = PatternFill("solid", fgColor=HEADER_COLOR)
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for col_idx, col_name in enumerate(df.columns, start=1):
            letter = get_column_letter(col_idx)
            max_len = max(
                len(str(col_name)),
                *[
                    len(str(ws.cell(row=r, column=col_idx).value or ""))
                    for r in range(2, ws.max_row + 1)
                ]
            )
            ws.column_dimensions[letter].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"

    def is_duplicate(self, invoice: Invoice) -> bool:
        """Check if invoice with same ID and vendor already exists."""
        if not invoice.invoice_id:
            return False

        inv_id     = str(invoice.invoice_id or "").strip().lower()
        inv_vendor = str(invoice.vendor_name or "").strip().lower()

        for excel_file in OUTPUTS_DIR.glob("*.xlsx"):
            try:
                existing_df = pd.read_excel(excel_file, sheet_name="Invoices")

                if "Invoice ID" not in existing_df.columns:
                    continue

                match = (
                    existing_df["Invoice ID"]
                    .astype(str).str.strip().str.lower() == inv_id
                ) & (
                    existing_df["Vendor"]
                    .astype(str).str.strip().str.lower() == inv_vendor
                )

                if match.any():
                    logger.warning(
                        f"Duplicate found in {excel_file.name}: "
                        f"invoice_id='{invoice.invoice_id}' "
                        f"vendor='{invoice.vendor_name}'"
                    )
                    return True

            except Exception as e:
                logger.warning(f"Cannot check {excel_file.name}: {e}")
                continue

        return False

    def move_to_valid_invoices(self, file_path: Path) -> None:
        """Copy validated invoice to /data/valid_invoices/."""
        from config.settings import VALID_INVOICES_DIR
        dest = VALID_INVOICES_DIR / file_path.name
        shutil.copy(str(file_path), str(dest))
        logger.info(f"Copied to valid_invoices: {file_path.name}")

    def move_to_need_review(self, file_path: Path) -> None:
        """Move failed invoice to /data/need_review/."""
        dest = NEED_REVIEW_DIR / file_path.name
        shutil.move(str(file_path), str(dest))
        logger.info(f"Moved to need_review: {file_path.name}")

    def cleanup_pending(self, file_path: Path) -> None:
        """Remove file from /data/pending/ after processing."""
        if file_path.exists():
            file_path.unlink()
            logger.debug(f"Removed from pending: {file_path.name}")